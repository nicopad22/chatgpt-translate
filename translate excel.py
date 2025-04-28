import openai
import openpyxl
import os
import time

# VARIABLES
root = "C:\\Users\\nicol\\Downloads\\CHATGPT-TRANSLATE\\Excel\\BHPF 1\\"
output_language = input("output language [es/en]: ")

# OpenAI API client
# Store there your API key
with open(".\\TOKEN.txt") as token_file:
    client = openai.OpenAI(api_key=[line for line in token_file][0])

#
# --- METHODS ---
#
def translate(text):
    if text.strip() != "":
        if output_language == "es":
            return translate_to_spanish(text)
        elif output_language == "en":
            return translate_to_english(text)
    else:
        return ""

# Function to translate text
def translate_to_spanish(text):
    result = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=[
                {"role": "system",
                 "content": "Eres un asistente de traducción. Traduces cualquier texto que recibas al español, sin realizar comentarios adicionales."},
                {"role": "user", "content": f"{text.strip()}"}
            ],
        temperature=0.7,
    )

    translation = result.choices[0].message.content
    print(text + " / " + translation)

    return translation

# Function to translate text
def translate_to_english(text):
    result = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=[
                {"role": "system",
                 "content": "You are a translation assistant. You translate any text into english, without making any comments even if it is not translatable."},
                {"role": "user", "content": f"{text.strip()}"}
            ],
        temperature=0.7,
    )

    translation = result.choices[0].message.content
    print(text + " / " + translation)

    return translation

# Applies given function to an entire sheet
def translate_sheet(input_workbook, out_workbook, sheet_name):
    sheet = input_workbook[sheet_name]
    output_sheet = out_workbook[sheet_name]

    # Iterate through each row and column
    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            value = cell.value

            # check if value is worth translating (string, not empty, not numeric, not formula)
            if isinstance(value, str):
                if not value.isnumeric() and value.strip() != "" and value[0] != "=":
                    
                    # update cell value
                    translated_value = translate(value)

                    # update the copied sheet
                    output_sheet.cell(row=row_idx, column=col_idx, value=translated_value)
                else:
                    output_sheet.cell(row=row_idx, column=col_idx, value=value)
            else:
                output_sheet.cell(row=row_idx, column=col_idx, value=value)


#
# --- MAIN ---
#

# List out all files in folder, filter out != .xlsx
files = os.listdir(root)
print(f"files found: ")
excel_files = []
for file in files:
    if file[-5:] == ".xlsx":
        print(" - " + file)
        excel_files.append(file)

# iterate all files
for file in excel_files:
    filepath = root + file
    out_filepath = filepath[:-5] + " [TRANSLATED].xlsx"

    if (file[:-5] + " [TRANSLATED].xlsx") in files:
        continue

    # make copy of original excel file to edit (keep formatting)
    os.popen(f"copy \"{filepath}\" \"{out_filepath}\"")
    time.sleep(0.5)

    # Load the workbook
    workbook = openpyxl.load_workbook(filepath)

    # Create a new workbook for the output
    output_workbook = openpyxl.load_workbook(out_filepath)

    # Iterate through each sheet
    for sheet_name in workbook.sheetnames:
        translate_sheet(workbook, output_workbook, sheet_name)
        output_workbook[sheet_name].title = translate(sheet_name)
        
    # Save the output workbook
    output_workbook.save(out_filepath)